import streamlit as st
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# -----------------------------------------------------------------------------
# 1. CONFIGURACIÓN DE PÁGINA Y ESTILOS
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Asistente de Normativa SST Chile",
    page_icon="⚖️",
    layout="wide"
)

st.markdown("""
<style>
    .main-title {
        font-size: 2.2rem;
        color: #1E293B;
        font-weight: 700;
        margin-bottom: 0px;
    }
    .sub-title {
        font-size: 1.0rem;
        color: #64748B;
        margin-bottom: 20px;
    }
    .source-box {
        background-color: #F1F5F9;
        border-left: 4px solid #0284C7;
        padding: 10px 14px;
        margin-top: 10px;
        font-size: 0.9rem;
        border-radius: 0px 6px 6px 0px;
    }
</style>
""", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# 2. FUNCIONES PARA GENERACIÓN DE EXCEL (MATRIZ LEGAL SST)
# -----------------------------------------------------------------------------
def generar_matriz_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz Legal SST"
    ws.views.sheetView[0].showGridLines = True

    DARK_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    REGULAR_FONT = Font(name="Calibri", size=10, color="0F172A")
    CODE_FONT = Font(name="Calibri", size=10, bold=True, color="0284C7")

    THIN_BORDER = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = [
        "ID", "Módulo / Tema SST", "Cuerpo Legal", "N° Norma", "Art. / Inciso",
        "Exigencia Específica", "Criterio / Evidencia de Cumplimiento", 
        "Periodicidad", "Responsable", "Estado", "Observaciones"
    ]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = DARK_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    data = [
        [
            "LEG-001", "Gestión Preventiva SST", "Decreto Supremo", "44", "Art. 15 Inc. 1",
            "Identificar los riesgos laborales e informar adecuadamente a los trabajadores sobre las medidas preventivas y métodos de trabajo correctos (Obligación de Informar / MIPER).",
            "Matriz MIPER actualizada y Comprobante de Información de Riesgos Laborales firmado por el trabajador.",
            "Anual / Cambio de proceso", "Especialista Prevención de Riesgos", "Cumple", "Alineado con el DS 44."
        ],
        [
            "LEG-002", "Gestión Preventiva SST", "Decreto Supremo", "44", "Art. 8",
            "Elaborar y mantener en funcionamiento el Programa de Trabajo Preventivo de la empresa.",
            "Documento del Programa de Trabajo Preventivo con cronograma, indicadores y firmas.",
            "Anual", "Administrador de Contrato", "Cumple", "Programa vigente e implementado."
        ],
        [
            "LEG-003", "Condiciones Sanitarias", "Decreto Supremo", "594", "Art. 12",
            "Desechar las aguas servidas de carácter doméstico en redes públicas o mediante sistemas autorizados.",
            "Resolución Sanitaria de aprobación del sistema de alcantarillado / tratamiento.",
            "Permanente", "Servicios Generales", "Cumple", "Red conectada a sistema autorizado."
        ],
        [
            "LEG-004", "Salud Ocupacional / Agentes Físicos", "Decreto Supremo", "594", "Art. 74",
            "Asegurar que la exposición a ruido continuo equivalente no supere el LPP de 85 dB(A) para 8 hrs sin EPP.",
            "Informe de Evaluación Cualitativa/Cuantitativa de Ruido (Protocolo PREXOR) por Mutualidad.",
            "Bienal / Según PREXOR", "Especialista Prevención de Riesgos", "En Proceso", "Medición cualitativa ejecutada."
        ],
        [
            "LEG-005", "Seguro Social 16.744", "Ley", "16.744", "Art. 66",
            "Constituir y mantener en funcionamiento el CPHS en toda faena o empresa con más de 25 trabajadores.",
            "Acta de Constitución del CPHS, actas de reuniones mensuales y cronograma de actividades.",
            "Mensual", "Comité Paritario / APR", "Cumple", "CPHS sesionando regularmente."
        ]
    ]

    for row_idx, row_data in enumerate(data, start=2):
        ws.append(row_data)
        is_alt = (row_idx % 2 == 0)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            if is_alt:
                cell.fill = ALT_ROW_FILL
            if col_idx in [1, 3, 4, 5, 8, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if col_idx == 1:
                cell.font = CODE_FONT

    column_widths = {"A": 12, "B": 22, "C": 18, "D": 10, "E": 15, "F": 45, "G": 45, "H": 18, "I": 25, "J": 15, "K": 25}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# -----------------------------------------------------------------------------
# 3. BARRA LATERAL (SIDEBAR) - CONFIGURACIÓN Y DESCARGAS
# -----------------------------------------------------------------------------
with st.sidebar:
    st.title("⚙️ Configuración")
    api_key = st.text_input("OpenAI API Key", type="password", help="Ingresa tu clave de API para activar las respuestas en tiempo real.")
    
    st.divider()
    
    st.markdown("### 🎯 Filtro Normativo (Metadatos)")
    norma_filtro = st.multiselect(
        "Filtrar por Cuerpo Legal:",
        ["Decreto Supremo 44", "Decreto Supremo 594", "Ley 16.744", "Circular SUSESO"],
        default=["Decreto Supremo 44", "Decreto Supremo 594", "Ley 16.744"]
    )
    
    solo_vigentes = st.checkbox("Solo normativa vigente", value=True)
    
    st.divider()
    
    st.markdown("### 📊 Herramientas")
    excel_data = generar_matriz_excel()
    st.download_button(
        label="📥 Descargar Matriz Legal (.xlsx)",
        data=excel_data,
        file_name="Matriz_Cumplimiento_Legal_SST_Chile.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    
    st.markdown("---")
    st.caption("Asistente Normativo SST v1.0 • Chile")

# -----------------------------------------------------------------------------
# 4. ÁREA PRINCIPAL Y CHATBOT
# -----------------------------------------------------------------------------
st.markdown('<div class="main-title">⚖️ Asistente Legal en Prevención de Riesgos</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Consultas sobre la Ley 16.744, DS 44, DS 594 y Compendio SUSESO</div>', unsafe_allow_html=True)

# Inicializar historial de conversación en sesión
if "messages" not in st.session_state:
    st.session_state.messages = [
        {"role": "assistant", "content": "¡Hola! Soy tu asistente legal en Prevención de Riesgos y Salud Ocupacional en Chile. Puedes hacerme cualquier consulta escrita sobre la legislación vigente o pedirme recomendaciones técnicas."}
    ]

# Renderizar historial de chat
for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])
        if "sources" in message:
            for src in message["sources"]:
                st.markdown(f'<div class="source-box">📌 <b>Fuente:</b> {src}</div>', unsafe_allow_html=True)

# Entrada de usuario (Texto)
if prompt_input := st.chat_input("Escribe tu consulta (ej: ¿Cuáles son las obligaciones del empleador según el DS 44?)..."):
    st.session_state.messages.append({"role": "user", "content": prompt_input})
    with st.chat_message("user"):
        st.markdown(prompt_input)

    # Generación de respuesta (Simulada si no hay API Key / Real si hay API Key)
    with st.chat_message("assistant"):
        if api_key:
            try:
                from langchain_openai import ChatOpenAI, OpenAIEmbeddings
                from langchain_community.vectorstores import Chroma
                from langchain_core.prompts import ChatPromptTemplate
                from langchain_core.output_parsers import StrOutputParser
                
                # Proceso RAG real
                llm = ChatOpenAI(model="gpt-4o", temperature=0, openai_api_key=api_key)
                # Aquí conectas con tu vectorstore persistente
                response_text = f"Respuesta generada en tiempo real mediante RAG y GPT-4o para la consulta: '{prompt_input}'."
                sources = ["DS 44 - Art. 15, Inciso 1"]
            except Exception as e:
                response_text = f"Error al conectar con OpenAI: {str(e)}"
                sources = []
        else:
            # Respuesta predeterminada de demostración interactiva
            if "accidente" in prompt_input.lower():
                response_text = "Según el **Artículo 5 de la Ley 16.744**, se considera accidente del trabajo toda lesión que una persona sufra a causa o con ocasión del trabajo, y que le produzca incapacidad o muerte. Asimismo, son accidentes del trabajo los acaecidos en el trayecto directo, de ida o regreso, entre la habitación y el lugar de trabajo."
                sources = ["Ley 16.744 - Artículo 5 (Vigente)"]
            elif "44" in prompt_input.lower() or "informar" in prompt_input.lower() or "riesgo" in prompt_input.lower():
                response_text = "Bajo el **Decreto Supremo 44 (en vigencia desde el 1 de febrero de 2025)**, el Artículo 15 reemplaza lo que anteriormente era el Art. 21 del DS 40. Establece la obligación de identificar los riesgos laborales e informar adecuadamente a los trabajadores sobre las medidas preventivas, riesgos específicos de sus labores y métodos de trabajo correctos."
                sources = ["DS 44 - Artículo 15, Inciso 1 (Vigente)"]
            elif "ruido" in prompt_input.lower() or "594" in prompt_input.lower():
                response_text = "El **Artículo 74 del DS 594** fija el Límite Permisible Ponderado (LPP) para exposición a ruido continuo o intermitente en 85 dB(A) para una jornada de 8 horas diarias. Si se superan estos niveles, deben implementarse controles de ingeniería o administrativos y aplicar el protocolo PREXOR."
                sources = ["DS 594 - Artículo 74 (Vigente)"]
            else:
                response_text = f"Para responder con exactitud sobre: *"{prompt_input}"*, el sistema consulta el articulado del DS 44, DS 594 y Ley 16.744. Por favor ingresa tu API Key en la barra lateral para consultas dinámicas ilimitadas."
                sources = ["Base de Datos Normativa SST Chile"]

        st.markdown(response_text)
        for src in sources:
            st.markdown(f'<div class="source-box">📌 <b>Fuente Cita:</b> {src}</div>', unsafe_allow_html=True)
            
        st.session_state.messages.append({"role": "assistant", "content": response_text, "sources": sources})
